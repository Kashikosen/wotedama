# -*- coding: utf-8 -*-
# kill -9 $(lsof -t -i:12345)
# wotedama_alpha.py

import MeCab
from transformers import AutoTokenizer, AutoModelForMaskedLM
import torch
import socket
import PySimpleGUI as sg
import threading
import statistics
import re
from openpyxl import load_workbook
import jaconv # type: ignore


# GPUの使用
device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")

# モデルとトークナイザーの読み込み
model_name = "cl-tohoku/bert-base-japanese-whole-word-masking"
tokenizer = AutoTokenizer.from_pretrained(model_name)
model = AutoModelForMaskedLM.from_pretrained(model_name)

# MeCabのセットアップ
tagger = MeCab.Tagger('-Owakati -d /opt/homebrew/lib/mecab/dic/mecab-ipadic-neologd')
yomi = MeCab.Tagger('-Oyomi -d /opt/homebrew/lib/mecab/dic/mecab-ipadic-neologd')

# 文字種判別
moji_file_pass = '/Users/Asuka/mysite/wotedama/moji.txt'
with open(moji_file_pass, 'r', encoding='UTF-8') as file:
    kanji = file.read()
moji = re.compile(f'[{kanji}]+')
hira = re.compile('[\u3041-\u309F]+')

# Excelのロード
wb = load_workbook('/Users/Asuka/mysite/wotedama/datasets_hinshi.xlsx', data_only=True)
ws = wb['Sheet1']   # Wikipedia頻度データ

# ソケット通信
M_SIZE = 1024
host = '127.0.0.1'
port = 12345

# ソケットを作成
sock = socket.socket(socket.AF_INET, type=socket.SOCK_STREAM)
print('create socket')

sock.bind((host, port))



class Context:

    def __init__(self):

        # テキストファイル
        self.file_pass = ''

        # 入力
        self.options = []
        self.typing_sentences = ''
        self.sentences = ''
        self.sentences_mask = ''
        self.meishi_list = []

        # 品詞予測
        self.hinshi_last = ''
        self.hinshi_prediction = {}

        threading.Thread(target=self.receive_messages_and_predict_conversions, daemon=True).start()


    ### 文字入力GUI ###
    def create_text_gui(self):
        # レイアウト
        layout = [
            [sg.FileBrowse('Open (⌘o)', key='bt_open', enable_events=True)],
            [sg.Multiline(default_text='', size=(150, 50),
                          border_width=2, key='text',
                          text_color='#000000', background_color='#ffffff')],
            [sg.Button('Read (⌘r)', key='bt_read'), sg.Button('Clear (⌘l)', key='bt_clear'), sg.Button('Save (⌘s)', key='bt_save'), sg.Button('Quit (⌘q)', key='bt_quit')]]

        # ウィンドウ作成
        window = sg.Window('wotedama_alpha.py', layout, return_keyboard_events=True)

        # イベントループ
        while True:
            # イベントの読み取り（イベント待ち）
            event, values = window.read()

            # Quit
            if event is None or event == 'bt_quit' or event == 'q' and ('Meta_L' or 'Meta_R'):
                break

            # Read
            elif event == 'bt_read' or event == 'r' and ('Meta_L' or 'Meta_R') or event.startswith('Return'):
                print(values['text'])
                self.sentences = values['text']

            # Clear
            elif event == 'bt_clear' or event == 'l' and ('Meta_L' or 'Meta_R'):
                window['text'].update('')
                self.sentences = ''
                self.meishi_list = []
            
            # Save
            elif event == 'bt_save' or event == 's' and ('Meta_L' or 'Meta_R'):
                # 上書き保存
                if self.file_pass:
                    with open(self.file_pass, 'w', encoding='UTF-8') as file:
                        file.write(values['text'])
                        print("上書き保存完了")
                else:
                    self.file_pass = sg.popup_get_file('save', save_as='True')
                    print(self.file_pass)
                    with open(self.file_pass, 'w', encoding='UTF-8') as file:
                        file.write(values['text'])
                        print("新規保存完了")
            
            # Open
            elif event == 'bt_open' or event == 'o' and ('Meta_L' or 'Meta_R'):
                self.file_pass = values['bt_open']

                if self.file_pass:
                    with open(self.file_pass, 'r', encoding='UTF-8') as file:
                        read_text = file.read()
                        window['text'].update(read_text)
                        print("ファイルを開く")

        # 終了表示
        window.close()

    ###



    ### 変換候補の受信 ###
    def receive_messages_and_predict_conversions(self):

        # 接続待機
        sock.listen(1)
        print('Waiting for connection')

        # 接続
        conn, cli_addr = sock.accept()
        print(f"Connection established with {cli_addr}")

        # バッファーのリセット
        buffer = ""

        while True:
            try:
                print('Waiting message')

                # メッセージ受信
                message = conn.recv(M_SIZE)
                if not message:
                    print('Connection closed by the client.')
                    break

                buffer += message.decode(encoding='UTF-16')

                # メッセージを行毎に処理
                while '\n' in buffer:
                    line, buffer = buffer.split('\n', 1)
                    line = line.strip()
                    print(f'Received message is [{line}]')

                    # メッセージ判定(入力文or変換候補)
                    if line.startswith("sentences"):
                        line = line[10:]
                        self.typing_sentences = line
                        self.typing_sentences = jaconv.hira2kata(self.typing_sentences)     # ひらがなからカタカナに変更
                        self.sentences_mask = [self.sentences + '[MASK]']
                        print(f"入力中文字: {self.typing_sentences}")
                        print(f"入力文: {self.sentences_mask}")

                    elif line.startswith("candidates"):
                        line = line[11:]
                        line = line.split(',')
                        self.options = line
                        print(f"変換候補: {self.options}")

                        # 予測処理呼び出し
                        print("予測処理を呼び出します")
                        self.predict_masked_word(conn)

            except Exception as error:
                print(f'Error occurred: {error}')
                break

            except KeyboardInterrupt:
                print('\nClosing connection')
                conn.close()
                break

    ###



    # 予測
    def predict_masked_word(self, conn):
        print("予測開始")

        # トークナイズ
        inputs = tokenizer(self.sentences_mask, return_tensors="pt")

        # 'token_type_ids' を含むキーを削除
        inputs.pop("token_type_ids", None)

        mask_token_index = torch.where(inputs["input_ids"] == tokenizer.mask_token_id)[1]
        if len(mask_token_index) == 0:
            print("マスクトークンが見つかりませんでした")
            return
        print(f"トークナイズされたデータ: {inputs['input_ids']}")

        # マスクされたトークンの予測
        with torch.no_grad():
            outputs = model(**inputs)

            logits = outputs.logits
            mask_token_logits = logits[0, mask_token_index, :]
            print(f"mask_token_logits: {mask_token_logits.size()}")

        # スコア算出を呼び出し
        scores = self.calculate_scores(mask_token_logits, conn)
        return scores


    # スコア算出
    def calculate_scores(self, mask_token_logits, conn):
        print("スコア算出")
        scores = {}
        for option in self.options:
            # トークナイズしてIDを取得
            option_id = tokenizer.convert_tokens_to_ids(tokenizer.tokenize(option))[0]
            if option_id >= mask_token_logits.size(1):
                print(f"無効なoption_id: {option_id}")
                continue
            # スコアを取得
            scores[option] = mask_token_logits[0, option_id].item()
        print(f"スコア: {scores}")

        # スコアが高い順に並び替える
        sorted_options = sorted(self.options, key=lambda opt: scores.get(opt, -float('inf')), reverse=True)
        print(f"ソート後: {sorted_options}")

        # スコアが最も高い選択肢を選ぶ
        best_option = max(scores, key=scores.get)
        print(f"最適な選択肢: {best_option}")

        # スコアの平均値と標準偏差を算出
        mean_scores = statistics.mean(scores.values())
        sd_scores = statistics.pstdev(scores.values())
        print(f"スコアの平均値: {mean_scores}")
        print(f"スコアの標準偏差: {sd_scores}")
        
        # 品詞予測実行
        if self.sentences != '':
            print("品詞予測実行")
            self.hinshi()



        # スコア修正

        # 特定文字種を持たない変換候補を修正 <スコア修正α>
        for option, score in scores.items():
            if moji.search(option):
                pass
            else:
                scores[option] = score - (sd_scores * 3)
                print(f"スコア修正α: {option}    {score} → {scores[option]}")
        

        # 読み方が一致していたら上方修正（全てひらがな以外） <スコア修正β>
        for option, score in scores.items():
            yomikata = yomi.parse(option)
            yomikata = yomikata.replace('\n', '')
            #print(yomikata)

            if hira.fullmatch(option):
                print("全てひらがな")

            else:
                if yomikata == self.typing_sentences:
                    scores[option] = score + (sd_scores * 3)
                    print(f"スコア修正β: {option}    {score} → {scores[option]}")

                    # 文中に存在する一般名詞・固有名詞に加点 <スコア修正β'>
                    wakati_option = tagger.parse(option)
                    wakati_option_list = wakati_option.split()
                    for option_list in wakati_option_list:
                        if option_list in self.meishi_list:
                            scores[option] = score + (sd_scores * 6)
                            print(f"スコア修正β': {option}    {score} → {scores[option]}")

                else:
                    #print("読み方一致せず")
                    pass

        
        # 各品詞の繋がりやすさに合わせてスコア修正 <スコア修正γ>
        if self.options != []:
            hinshi_options_dict = {}

            for option, score in scores.items():
                node = tagger.parseToNode(option)

                for i in range(2):
                    if node.stat != 2 and node.stat != 3:
                        features = node.feature.split(',')
                        hinshi_kihon = features[0] if len(features) > 0 else ''
                        hinshi_shosai = features[1] if len(features) > 1 else '*'
                        
                        if hinshi_kihon:
                            hinshi_options_dict[option] = hinshi_kihon + '・' + hinshi_shosai
                            
                    node = node.next

                hinshi_probability = self.hinshi_prediction.get(hinshi_options_dict.get(option, ""), 0)
                scores[option] = score + (3 * hinshi_probability)
                print(f"スコア修正γ: {option}    {score} → {scores[option]}")

            print(hinshi_options_dict)



        # ## 没スコア修正案

        # # スコアが標準偏差より高いものは後ろにする
        # for option, score in scores.items():
        #     if score >= sd_scores:
        #         scores[option] = score - (sd_scores * 3)
        #         print(f"スコア修正：{option}  {score}→{scores[option]}")

        # # 最大スコアの文字のスコアを修正
        # for option, score in scores.items():
        #     if score == max(scores.values()):
        #         scores[option] = score - (sd_scores * 3)
        #         print(f"スコア修正1: {option}    {score} → {scores[option]}")

        # for option, score in scores.items():
        #     if len(option) >= 2:
        #         if moji.search(option):
        #             scores[option] = score + (len(option)*2)
        #             print(f"スコア修正2: {option}    {score} → {scores[option]}")
        #     else:
        #         scores[option] = score - 3
        #         print(f"スコア修正3: {option}    {score} → {scores[option]}")
        


        new_sorted_options = sorted(self.options, key=lambda opt: scores.get(opt, -float('inf')), reverse=True)
        print(f"スコア修正後ソート: {new_sorted_options}")


        # 解析結果を送信
        self.send_candidates_message(conn, new_sorted_options)


    ### 入力済み文章 ###
    # 基本品詞・詳細品詞表示
    def hinshi(self):

        hinshi_result_list = []

        # 空白や改行を消去
        cleaned_sentence = re.sub(r'\s+', '', self.sentences.strip())
        node = tagger.parseToNode(cleaned_sentence)

        # 基本品詞と詳細品詞の取得
        while node:
            if node.stat == 2 or node.stat == 3:
                node = node.next
                continue

            features = node.feature.split(',')
            hinshi_kihon = features[0] if len(features) > 0 else ''
            hinshi_shosai = features[1] if len(features) > 1 else '*'

            if hinshi_kihon:
                hinshi_result_list.append(f'{hinshi_kihon}・{hinshi_shosai}')
                #print(hinshi_result_list)

                # 名詞リストの更新
                if f'{hinshi_kihon}・{hinshi_shosai}' in ['名詞・一般','名詞・固有名詞']:
                    if node.surface not in self.meishi_list:
                        self.meishi_list.append(node.surface)

            node = node.next

        print(self.meishi_list)


        # 最後の単語の品詞を取得
        if hinshi_result_list:
            self.hinshi_last = hinshi_result_list[-1] if hinshi_result_list else 'なし'
            print(f"品詞リスト:  {hinshi_result_list}")
            print(f"最後の単語の品詞: {self.hinshi_last}")
        else:
            self.hinshi_last = '※対象が存在しません'

        # BOS/EOS・*を記号・句点扱いとして処理
        if self.hinshi_last == 'BOS/EOS・*':
            self.hinshi_last = '記号・句点'

        # 品詞予測実行
        self.predict_hinshi()
    

    # 品詞予測
    def predict_hinshi(self):

        # 最後の単語の品詞に対応する行を検索
        row = None
        for r in range(2, ws.max_row + 1):
            hinshi_location = ws[f'A{r}'].value
            if hinshi_location == self.hinshi_last:
                row = r
                break
        # print(f"{row}行目")

        # 見つからなかった場合
        if row is None:
            print(f"Error: {self.hinshi_last}に対応する行が見つかりませんでした")
            return
        
        # 予測処理
        total_cell = ws[f'AY{row}'].internal_value  # n単語目の合計件数
        # print(f"{total_cell}件")

        if total_cell:
            total = int(total_cell) + 49    # 合計を整数に変換（0件が発生しないように+49）

            # n+1単語目の品詞毎に処理
            predictions = {}
            for col in range(2, ws.max_column):
                hinshi_next = ws.cell(row=1, column=col).value  # n+1文節目の品詞
                count_cell = ws.cell(row=row, column=col).value     # 件数
                # print(hinshi_next, count_cell)

                if count_cell:
                    try:
                        count = int(count_cell) + 1     # 件数を整数に変換（0件が発生しないように+1）
                    except ValueError:
                        print(f"Error: {count_cell}")
                        continue
                    
                    # 頻度から確率を計算
                    if count > 0:
                        probability = count / total
                        predictions[hinshi_next] = probability
                        # print(predictions[hinshi_next])
            
            # 確率が高い順に並び替える
            sorted_hinshi_prediction = sorted(predictions.items(), key=lambda x: x[1], reverse=True)

            # 辞書型で保持
            self.hinshi_prediction = {hinshi: probability for hinshi, probability in sorted_hinshi_prediction}

            # 結果を表示
            prediction_result = "\n".join([f"{hinshi}: {probability:.2%}" for hinshi, probability in sorted_hinshi_prediction])
            print(prediction_result)
        
        ###
    

    # 解析結果を送信
    def send_candidates_message(self, conn, new_sorted_options):
        print('Send response to Client')
        sorted_options_str = str(new_sorted_options)
        conn.send(sorted_options_str.encode(encoding='UTF-16'))
        print(f'Send message is [{sorted_options_str}]')
        print("-" * 50)

###

context = Context()
context.create_text_gui()
